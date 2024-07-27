from openpyxl import load_workbook
import re
from openpyxl.utils import get_column_letter
import xlrd


def col_index_from_string(col_str):
    col_index = 0
    for char in col_str:
        col_index = col_index * 26 + (ord(char) - ord('A') + 1)
    return col_index - 1  # 减1是因为索引通常从0开始


# rules def [("a1","b1","fdasfda"),("a2","b2"),"fdaf"]


def extract_cell_coordinates(cell_ref):
    """
    Extract the column and row numbers from an Excel cell reference.

    :param cell_ref: str, the cell reference like 'A10', 'B2', 'Z10', etc.
    :return: tuple, (column, row) as integers
    """
    # Regular expression pattern for matching Excel cell references
    pattern = re.compile(r'^([A-Za-z]+)(\d+)$')

    match = pattern.match(cell_ref)
    if not match:
        raise ValueError(f"Invalid cell reference format: {cell_ref}")

    # Extract column letters and row numbers
    column_letters = match.group(1).upper()
    row_number = int(match.group(2))
    return column_letters, row_number


def get_data_xldr(ws, row, col):
    return ws.cell_value(row-1, col_index_from_string(col))


def copy_data(ss, ds, rules, lambda_col=None, lambda_row=None, strict=True, xlrd=True):
    for rule in rules:
        s_c, s_r = extract_cell_coordinates(rule[0])
        s_v = get_data_xldr(ss, s_r, s_c)
        print(s_c, s_r, s_v)
        d_c, d_r = extract_cell_coordinates(rule[1])

        if lambda_col:
            d_c = lambda_col(d_c)
        ds["{}{}".format(d_c, d_r)].value = s_v


def copy_zichanfuzhai(src_filename, dst_filename, month):
    rules = [
        ("c6", "d98", ""),
    ]
    src_wb = xlrd.open_workbook(filename=src_filename)
    src_ws = src_wb.sheet_by_index(0)
    dst_wb = load_workbook(filename=dst_filename)
    dst_ws = dst_wb["资产负债表"]

    def cal_month_index(row):
        assert (len(row) == 1)
        # 资产负债表中月份的索引为 "D"
        return str(ord("D") + month-1)

    copy_data(src_ws, dst_ws, rules, lambda_row=cal_month_index)
    dst_wb.save(dst_filename)
    dst_wb.close()


if __name__ == '__main__':
    src_filename = "/home/bhzk/abc/202405/源数据/北京轻舟创游科技有限公司北京海淀分公司_资产负债表_202405.xls"
    dst_filename = "/home/bhzk/abc/202405/轻舟创游总分支机构表-202405 - 无链接.xlsx"
    copy_zichanfuzhai(src_filename, dst_filename, 1)
